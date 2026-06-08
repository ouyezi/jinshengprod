import io
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models import EMPLOYEE_LEVELS, UserInfo, next_target_level
from app.pinyin_util import name_to_pinyin_keys

TEMPLATE_HEADERS = [
    "分管中心", "一级部门", "工号", "姓名", "学历", "岗位", "职级",
    "FY24年度等级", "FY25年度等级", "FY25H1等级", "入职时间", "备注", "提名情况", "提名理由",
]
SUMMARY_HEADERS = [
    "评审对象姓名", "状态", "修改时间", "目标职级", "评委姓名", "评审日期",
    "价值观平均分", "能力模型平均分", "工作成果平均分", "最终总分",
    "系统建议", "评委确认结果",
    "务实评分", "担当评分", "追求卓越评分",
    "学习创新与效率提升评分", "技术专业与质量评分", "架构能力评分",
    "业务理解能力评分", "执行力评分", "团队协作评分", "知识传承与影响力评分",
    "基础工作产出评分", "AI使用深度评分",
    "突出优势", "待发展项",
]


def _cell_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _parse_join_date(val) -> Optional[date]:
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val).strip()[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "员工信息"
    ws.append(TEMPLATE_HEADERS)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_employees(db: Session, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    success, errors = 0, []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        employee_no = _cell_str(row[2] if len(row) > 2 else None)
        if not employee_no:
            continue

        name = _cell_str(row[3] if len(row) > 3 else None)
        current_level = _cell_str(row[6] if len(row) > 6 else None)

        if not name:
            errors.append({"row": idx, "reason": "姓名为空"})
            continue
        if not current_level or current_level not in EMPLOYEE_LEVELS:
            errors.append({"row": idx, "reason": f"职级无效: {current_level}"})
            continue

        target_level = next_target_level(current_level)
        if not target_level:
            errors.append({"row": idx, "reason": f"无法自动推算目标职级: {current_level}"})
            continue

        fields = {
            "employee_no": employee_no,
            "name": name,
            "name_pinyin": name_to_pinyin_keys(name),
            "division_center": _cell_str(row[0] if len(row) > 0 else None),
            "department": _cell_str(row[1] if len(row) > 1 else None),
            "education": _cell_str(row[4] if len(row) > 4 else None),
            "position": _cell_str(row[5] if len(row) > 5 else None),
            "current_level": current_level,
            "target_level": target_level,
            "perf_fy24": _cell_str(row[7] if len(row) > 7 else None),
            "perf_fy25": _cell_str(row[8] if len(row) > 8 else None),
            "perf_fy25h1": _cell_str(row[9] if len(row) > 9 else None),
            "join_date": _parse_join_date(row[10] if len(row) > 10 else None),
            "remark": _cell_str(row[11] if len(row) > 11 else None),
            "nomination_status": _cell_str(row[12] if len(row) > 12 else None),
            "nomination_reason": _cell_str(row[13] if len(row) > 13 else None),
            "update_time": datetime.utcnow(),
        }

        existing = db.query(UserInfo).filter(UserInfo.employee_no == employee_no).first()
        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
        else:
            db.add(UserInfo(**fields))
        success += 1

    db.commit()
    return {"success": success, "errors": errors}


def build_summary_export(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(SUMMARY_HEADERS)
    for r in rows:
        ws.append([r.get(h) for h in SUMMARY_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
